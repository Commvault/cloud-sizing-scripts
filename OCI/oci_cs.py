import sys
import subprocess
import oci
from oci.monitoring.models import SummarizeMetricsDataDetails
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import os
from datetime import datetime
import logging

total_instances = 0
total_instance_sizeGB = 0
total_instance_sizeTB = 0
total_namespaces = 0
total_buckets = 0
total_storageGB = 0
total_storageTB = 0

class InstanceInfo:
    def __init__(self):
        self.compartment_id = None
        self.instance_id = None
        self.instance_name = None
        self.region = None
        self.availability_domain = None
        self.shape = None
        self.state = None
        self.number_of_volumes = 0
        self.sizeGB = 0
        self.sizeTB = 0
        self.defined_tags = {}
        self.freeform_tags = {}

class InstanceSummary:
    def __init__(self):
        self.region = None
        self.compartment_id = None
        self.instance_count = 0
        self.total_sizeGB = 0
        self.total_sizeTB = 0

class ObjectStorageInfo:
    def __init__(self):
        self.compartment_id = None
        self.namespace = None
        self.bucket_name = None
        self.region = None
        self.storage_tier = None
        self.object_count = 0
        self.sizeGB = 0
        self.sizeTB = 0
        self.defined_tags = {}
        self.freeform_tags = {}

class ObjectStorageSummary:
    def __init__(self):
        self.region = None
        self.namespace = None
        self.compartment_id = None
        self.bucket_count = 0
        self.total_storage_GB = 0
        self.total_storage_TB = 0

def install_and_import(package):
    try:
        __import__(package)
        print(f"Package '{package}' is already installed.")
    except ImportError:
        print(f"Package '{package}' not found. Installing...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

def get_sheet_info(workload):
    if workload == "instances":
        info_sheet = "InstanceInfo"
        summary_sheet = "InstanceSummary"
        info_headers = [
            "Compartment ID", "Instance ID", "Instance Name", "Region",
            "Availability Domain", "Shape", "State", "Number of Volumes",
            "Size (GB)", "Size (TB)", "Defined Tags", "Freeform Tags"
        ]
        summary_headers = ["Region", "Compartment ID", "Instance Count", "Total Size (GB)", "Total Size (TB)"]
    elif workload == "object_storage":
        info_sheet = "ObjectStorageInfo"
        summary_sheet = "ObjectStorageSummary"
        info_headers = [
            "Namespace", "Compartment ID", "Bucket Name", "Region",
            "Storage Tier", "Object Count", "Size (GB)", "Size (TB)",
            "Defined Tags", "Freeform Tags"
        ]
        summary_headers = ["Namespace", "Region", "Compartment ID", "Bucket Count", "Total Size (GB)", "Total Size (TB)"]
    else:
        raise ValueError(f"Unsupported workload: {workload}")

    return info_sheet, summary_sheet, info_headers, summary_headers

def format_workbook(filename):
    wb = load_workbook(filename)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    bold_font = Font(bold=True)

    for sheet in wb.worksheets:
        # Format header row
        for cell in sheet[1]:
            cell.font = bold_font
            cell.fill = header_fill

        # Adjust column widths
        for column_cells in sheet.columns:
            max_length = 0
            col = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col].width = adjusted_width
    wb.save(filename)

def init_excel(filename, workload):
    if not os.path.exists(filename):
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        info_sheet, summary_sheet, info_headers, summary_headers = get_sheet_info(workload)

        wb.create_sheet(info_sheet)
        wb.create_sheet(summary_sheet)

        wb[info_sheet].append(info_headers)
        wb[summary_sheet].append(summary_headers)
        wb.save(filename)

def dump_info(filename, workload, object_list: list):
    info_sheet, _, _, _ = get_sheet_info(workload)
    wb = load_workbook(filename)
    sheet = wb[info_sheet]

    for obj in object_list:
        # For now, assuming instances workload — extend this when you add object_storage
        if workload == "instances":
            row = [
                obj.compartment_id,
                obj.instance_id,
                obj.instance_name,
                obj.region,
                obj.availability_domain,
                obj.shape,
                obj.state,
                obj.number_of_volumes,
                obj.sizeGB,
                obj.sizeTB,
                str(obj.defined_tags),
                str(obj.freeform_tags),
            ]
        elif workload == "object_storage":
            row = [
                obj.namespace,
                obj.compartment_id,
                obj.bucket_name,
                obj.region,
                obj.storage_tier,
                obj.object_count,
                obj.sizeGB,
                obj.sizeTB,
                str(obj.defined_tags),
                str(obj.freeform_tags),
            ]
        else:
            raise ValueError(f"Unsupported workload: {workload}")
        sheet.append(row)
    wb.save(filename)

def dump_summary(filename, workload, summary):
    _, summary_sheet, _, _ = get_sheet_info(workload)
    wb = load_workbook(filename)
    sheet = wb[summary_sheet]

    if workload == "instances":
        row = [
            summary.region,
            summary.compartment_id,
            summary.instance_count,
            summary.total_sizeGB,
            summary.total_sizeTB,
        ]
    elif workload == "object_storage":
        row = [
            summary.namespace,
            summary.region,
            summary.compartment_id,
            summary.bucket_count,
            summary.total_storage_GB,
            summary.total_storage_TB,
        ]
    else:
        raise ValueError(f"Unsupported workload: {workload}")

    sheet.append(row)
    wb.save(filename)

def write_grand_total(filename, workload):
    wb = load_workbook(filename)

    if workload == "instances":
        global total_instances, total_instance_sizeGB, total_instance_sizeTB
        sheet_name = "InstanceSummary"
        row = ["Grand Total", "", total_instances, total_instance_sizeGB, total_instance_sizeTB]
    elif workload == "object_storage":
        sheet_name = "ObjectStorageSummary"
        # You'd update these totals accordingly if you track them
        # For example, let's assume:
        global total_namespaces, total_buckets, total_storageGB, total_storageTB
        row = ["Grand Total", "", total_buckets, total_storageGB, total_storageTB]
    else:
        raise ValueError(f"Unsupported workload: {workload}")

    sheet = wb[sheet_name]
    sheet.append(row)
    wb.save(filename)

def get_object_storage_info(config, filename, regions=[], compartments=[]):
    global total_buckets, total_storageGB, total_storageTB
    object_storage_summary_list = []
    identity_client = oci.identity.IdentityClient(config)
    # Get subscribed regions if regions list is empty
    if not regions:
        regions = [region.region_name for region in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [compartment.id for compartment in identity_client.list_compartments(compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]
    for region in regions:
        logging.info(f"Processing region: {region}")
        config["region"] = region
        object_storage_client = oci.object_storage.ObjectStorageClient(config)
        try:
            namespace = object_storage_client.get_namespace().data
        except Exception as e:
            logging.error(f"Error fetching namespace for region {region}: {e}")
            continue
        for compartment in compartments:
            compartment_summary = ObjectStorageSummary()
            compartment_summary.region = region
            compartment_summary.namespace = namespace
            compartment_summary.compartment_id = compartment
            compartment_summary.bucket_count = 0
            compartment_summary.total_storage_GB = 0
            compartment_summary.total_storage_TB = 0
            compartment_bucket_list = []
            try:
                buckets = oci.pagination.list_call_get_all_results(
                    object_storage_client.list_buckets,
                    namespace_name=namespace,
                    compartment_id=compartment
                ).data
            except Exception as e:
                logging.error(f"Error fetching buckets for compartment {compartment}: {e}")
                continue
            logging.info(f"Found {len(buckets)} bucket(s) in compartment {compartment}")
            if len(buckets) == 0:
                continue
            for bucket in buckets:
                bucket_info = ObjectStorageInfo()
                bucket_info.compartment_id = compartment
                bucket_info.namespace = namespace
                bucket_info.bucket_name = bucket.name
                bucket_info.region = region
                try:
                    stats = object_storage_client.get_bucket(
                        namespace_name=namespace,
                        bucket_name=bucket.name,
                        fields=['approximateSize', 'approximateCount']
                    ).data
                    bucket_info.storage_tier = stats.storage_tier
                    bucket_info.defined_tags = stats.defined_tags
                    bucket_info.freeform_tags = stats.freeform_tags
                    size_in_bytes, object_count = stats.approximate_size, stats.approximate_count
                    bucket_info.sizeGB = round(size_in_bytes / (1024 ** 3), 2) if size_in_bytes else 0
                    bucket_info.sizeTB = round(bucket_info.sizeGB / 1024, 2) if bucket_info.sizeGB else 0
                    bucket_info.object_count = object_count
                except Exception as e:
                    logging.error(f"Error fetching stats for bucket {bucket.name}: {e}")
                    continue
                compartment_summary.bucket_count += 1
                compartment_summary.total_storage_GB += bucket_info.sizeGB if bucket_info.sizeGB else 0
                compartment_summary.total_storage_TB += bucket_info.sizeTB if bucket_info.sizeTB else 0
                compartment_bucket_list.append(bucket_info)
                total_buckets += 1
                total_storageGB += bucket_info.sizeGB if bucket_info.sizeGB else 0
                total_storageTB += bucket_info.sizeTB if bucket_info.sizeTB else 0
            object_storage_summary_list.append(compartment_summary)
            dump_info(filename, "object_storage", compartment_bucket_list)
            dump_summary(filename, "object_storage", compartment_summary)
    write_grand_total(filename, "object_storage")
    format_workbook(filename)
    logging.info("Completed processing all regions and compartments for object storage.")
    logging.info(f"Grand Total - Buckets: {total_buckets}, Size (GB): {total_storageGB}, Size (TB): {total_storageTB}")

def get_boot_volume_size(config, instance_id, availability_domain, compartment_id):
    compute_client = oci.core.ComputeClient(config)
    block_storage_client = oci.core.BlockstorageClient(config)
    try:
        response = oci.pagination.list_call_get_all_results(compute_client.list_boot_volume_attachments,
                                                            instance_id=instance_id,
                                                            availability_domain=availability_domain,
                                                            compartment_id=compartment_id,
                                                            retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY)
        boot_volumes = response.data
        if not boot_volumes:
            return 0
        try:
            response = block_storage_client.get_boot_volume(boot_volumes[0].boot_volume_id)
            boot_volume_info = response.data
            return boot_volume_info.size_in_gbs
        except Exception as e:
            print(f"Error retrieving boot volume info for instance {instance_id}: {e}")
            return 0
    except Exception as e:
        print(f"Error retrieving boot volume attachments for instance {instance_id}: {e}")
        return 0

def get_block_volume_count_and_size(config, instance_id, availability_domain, compartment_id):
    compute_client = oci.core.ComputeClient(config)
    block_storage_client = oci.core.BlockstorageClient(config)
    try:
        response = oci.pagination.list_call_get_all_results(compute_client.list_volume_attachments,
                                                            instance_id=instance_id,
                                                            availability_domain=availability_domain,
                                                            compartment_id=compartment_id,
                                                            retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY)
        volume_attachments = response.data
        if not volume_attachments:
            return 0, 0
        total_size = 0
        for attachment in volume_attachments:
            try:
                response = block_storage_client.get_volume(attachment.volume_id)
                volume_info = response.data
                total_size += volume_info.size_in_gbs
            except Exception as e:
                print(f"Error retrieving volume info for volume {attachment.volume_id}: {e}")
        return len(volume_attachments), total_size
    except Exception as e:
        print(f"Error retrieving volume attachments for instance {instance_id}: {e}")
        return 0, 0

def get_instance_info(config, filename, regions=[], compartments=[]):
    global total_instances, total_instance_sizeGB, total_instance_sizeTB
    instance_summary_list = []
    identity_client = oci.identity.IdentityClient(config)
    # Get subscribed regions if regions list is empty
    if not regions: 
        regions = [region.region_name for region in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [compartment.id for compartment in identity_client.list_compartments(compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]
    for region in regions:
        logging.info(f"Processing region: {region}")
        config["region"] = region
        compute_client = oci.core.ComputeClient(config)
        for compartment in compartments:
            compartment_summary = InstanceSummary()
            logging.info(f"Processing compartment: {compartment}")
            compartment_instance_list = []
            compartment_summary.compartment_id = compartment
            compartment_summary.region = region
            instances = oci.pagination.list_call_get_all_results(compute_client.list_instances,
                                                                compartment_id=compartment,
                                                                retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY).data
            logging.info(f"Found {len(instances)} instance(s)")
            if len(instances) == 0:
                continue
            for instance in instances:
                if instance.lifecycle_state == "TERMINATED":
                    continue
                logging.info(f"Processing instance: {instance.id} - {instance.display_name}")
                instance_info = InstanceInfo()
                instance_info.compartment_id = compartment
                instance_info.instance_id = instance.id
                instance_info.instance_name = instance.display_name
                instance_info.region = region
                instance_info.availability_domain = instance.availability_domain
                instance_info.shape = instance.shape
                instance_info.state = instance.lifecycle_state
                instance_info.defined_tags = instance.defined_tags
                instance_info.freeform_tags = instance.freeform_tags
                try:
                    boot_volume_size = get_boot_volume_size(config, instance.id, instance.availability_domain, instance.compartment_id)
                    block_volume_count, block_volume_size = get_block_volume_count_and_size(config, instance.id, instance.availability_domain, instance.compartment_id)
                except Exception as e:
                    logging.error(f"Error fetching volume data for instance {instance.id}: {e}")
                    continue 
                instance_info.number_of_volumes = (1 if boot_volume_size > 0 else 0) + block_volume_count
                instance_info.sizeGB = boot_volume_size + block_volume_size 
                instance_info.sizeTB = round(instance_info.sizeGB / 1024, 2)
                compartment_summary.instance_count += 1
                compartment_summary.total_sizeGB += instance_info.sizeGB
                compartment_summary.total_sizeTB += instance_info.sizeTB
                compartment_instance_list.append(instance_info)
                total_instances += 1
                total_instance_sizeGB += instance_info.sizeGB
                total_instance_sizeTB += instance_info.sizeTB
            instance_summary_list.append(compartment_summary)
            dump_info(filename, "instances", compartment_instance_list)
            dump_summary(filename, "instances", compartment_summary)
    write_grand_total(filename, "instances")
    format_workbook(filename)
    logging.info("Completed processing all regions and compartments.")
    logging.info(f"Grand Total - Instances: {total_instances}, Size (GB): {total_instance_sizeGB}, Size (TB): {total_instance_sizeTB}")

if __name__ == "__main__":

    packages = ["oci", "openpyxl", "pandas"]

    for pkg in packages:
        install_and_import(pkg)

    args = sys.argv[1:]
    for arg in args:
        if arg.startswith("--profile="):
            profile_name = arg.split("=")[1]
        elif arg.startswith("--region="):
            regions = arg.split("=")[1].split(",")
        elif arg.startswith("--compartment="):
            compartments = arg.split("=")[1].split(",")
        elif arg.startswith("--workload="):
            workload = arg.split("=")[1]
        elif arg == "--help":
            print("Usage: python oci_cs.py --workload=<instances|object_storage> [--profile=<profilename>] [--region=<region1>,<region2>] [--compartment=<comp1>,<comp2>] [--help]")
            sys.exit(0)
        else:
            print(f"Unknown argument: {arg}")
            sys.exit(1)

    if "workload" not in locals():
        print("Error: --workload is required. Use --help for usage.")
        sys.exit(1)
    if "profile_name" not in locals():
        profile_name = oci.config.DEFAULT_PROFILE
    if "regions" not in locals():
        regions = []
    if "compartments" not in locals():
        compartments = []
    config = oci.config.from_file(profile_name=profile_name)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    log_dir = "Logs"
    os.makedirs(log_dir, exist_ok=True)
    log_filename = os.path.join(log_dir, f"{profile_name}_{workload}_{timestamp}.log")
    handlers=[
        logging.FileHandler(log_filename),
        logging.StreamHandler(sys.stdout)
    ]
    logging.basicConfig(
        handlers=handlers,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    metrics_dir = "Metrics"
    os.makedirs(metrics_dir, exist_ok=True)
    filename = os.path.join(metrics_dir, f"{profile_name}_{workload}_{timestamp}.xlsx")
    if workload == "instances":
        init_excel(filename,workload)
        get_instance_info(config, filename, regions, compartments)
    elif workload == "object_storage":
        init_excel(filename,workload)
        get_object_storage_info(config, filename, regions, compartments)
    else:
        print(f"Workload '{workload}' is not supported yet. Supported workloads: instances, object_storage")
        sys.exit(1)
